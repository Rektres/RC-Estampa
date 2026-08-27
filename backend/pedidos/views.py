import logging
from django.db import models
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from rest_framework import generics, mixins, permissions, status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from django.http import HttpResponse

from .models import Carrito, Cotizacion, ItemPedido, Pedido
from .serializers import CarritoSerializer, CotizacionSerializer, PedidoSerializer
from .services.mercadopago import obtener_info_pago_mercadopago

logger = logging.getLogger(__name__)


class PedidoViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    viewsets.GenericViewSet
):
    serializer_class = PedidoSerializer
    permission_classes = [permissions.AllowAny]
    lookup_field = 'numero'

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if getattr(user, 'rol', '') == 'admin' or user.is_staff:
                return Pedido.objects.all().order_by('-creado_en')
            return Pedido.objects.filter(models.Q(user=user) | models.Q(email__iexact=user.email)).order_by('-creado_en')
        email = self.request.query_params.get('email')
        if email:
            return Pedido.objects.filter(email__iexact=email).order_by('-creado_en')
        return Pedido.objects.none()

    @action(detail=True, methods=['patch', 'post'], permission_classes=[permissions.IsAuthenticated])
    def cambiar_estado(self, request, numero=None):
        from config.emails import enviar_email_cambio_estado
        user = request.user
        if getattr(user, 'rol', '') != 'admin' and not user.is_staff and not user.is_superuser:
            return Response(
                {"detail": "Solo administradores pueden modificar el estado del pedido."},
                status=status.HTTP_403_FORBIDDEN
            )

        pedido = self.get_object()
        nuevo_estado = request.data.get('estado')
        nota = request.data.get('nota', '').strip()
        if not nuevo_estado or nuevo_estado not in dict(Pedido.ESTADOS):
            return Response(
                {"detail": f"Estado inválido. Opciones: {list(dict(Pedido.ESTADOS).keys())}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        estado_anterior = pedido.estado
        pedido.estado = nuevo_estado
        if nuevo_estado == 'pagado' and not pedido.pagado_en:
            pedido.pagado_en = timezone.now()

        # Registrar en el historial de estados
        historial = list(pedido.historial_estados or [])
        historial.append({
            'estado_anterior': estado_anterior,
            'estado_nuevo': nuevo_estado,
            'fecha': timezone.now().isoformat(),
            'autor': getattr(user, 'nombre', '') or user.email,
            'nota': nota,
        })
        pedido.historial_estados = historial
        pedido.save(update_fields=['estado', 'pagado_en', 'historial_estados'])

        # Enviar notificación por correo al cliente
        enviar_email_cambio_estado(pedido, nuevo_estado, nota=nota)

        return Response(PedidoSerializer(pedido).data, status=status.HTTP_200_OK)



class CotizacionViewSet(mixins.CreateModelMixin, viewsets.GenericViewSet):
    serializer_class = CotizacionSerializer
    permission_classes = [permissions.AllowAny]
    queryset = Cotizacion.objects.all()


class CarritoView(generics.GenericAPIView):
    serializer_class = CarritoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_carrito(self):
        carrito, _ = Carrito.objects.get_or_create(user=self.request.user)
        return carrito

    def get(self, request):
        return Response(self.get_serializer(self.get_carrito()).data)

    def put(self, request):
        serializer = self.get_serializer(self.get_carrito(), data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


@csrf_exempt
@api_view(['POST', 'GET'])
@permission_classes([permissions.AllowAny])
def mercadopago_webhook(request):
    """
    Webhook / Notificación IPN para Mercado Pago.
    Recibe notificaciones de eventos de pago y actualiza el pedido correspondiente.
    """
    try:
        data = request.data if request.method == 'POST' else {}
        topic = request.GET.get('topic') or data.get('type')
        payment_id = request.GET.get('id') or data.get('data', {}).get('id') or request.GET.get('data.id')

        logger.info(f"Webhook Mercado Pago recibido: topic={topic}, payment_id={payment_id}")

        if (topic in ('payment', 'merchant_order') or data.get('action') == 'payment.created' or data.get('action') == 'payment.updated') and payment_id:
            payment_info = obtener_info_pago_mercadopago(payment_id)
            if payment_info:
                payment_status = payment_info.get("status")
                external_ref = payment_info.get("external_reference")

                logger.info(f"Estado de pago MP {payment_id} para ref {external_ref}: {payment_status}")

                if external_ref:
                    pedido = Pedido.objects.filter(numero=external_ref).first()
                    if pedido:
                        from .services.mercadopago_api import descontar_stock_y_notificar, poblar_datos_auditoria_pedido

                        pedido.transaccion_id = str(payment_id)
                        pedido.datos_pago_raw = payment_info

                        poblar_datos_auditoria_pedido(pedido, res=payment_info)

                        if payment_status == 'approved':
                            pedido.estado = 'pagado'
                            if not pedido.pagado_en:
                                pedido.pagado_en = timezone.now()
                            descontar_stock_y_notificar(pedido)
                        elif payment_status in ('rejected', 'cancelled'):
                            if pedido.estado == 'pendiente':
                                pedido.estado = 'cancelado'

                        pedido.save()
                        logger.info(f"Pedido {pedido.numero} actualizado exitosamente a estado={pedido.estado}")

        return Response({"status": "received"}, status=status.HTTP_200_OK)

    except Exception as exc:
        logger.error(f"Error procesando webhook de Mercado Pago: {exc}")
        return Response({"status": "error", "message": str(exc)}, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def procesar_pago_tarjeta(request):
    """
    Endpoint para procesar pagos con tarjeta directamente en el sitio (Checkout API).
    Crea el pedido si no existe y procesa el token de tarjeta con Mercado Pago.
    """
    try:
        from .services.mercadopago_api import procesar_pago_directo_mercadopago

        data = request.data
        token = data.get('token')
        payment_method_id = data.get('payment_method_id')
        installments = data.get('installments', 1)
        issuer_id = data.get('issuer_id')
        doc_type = data.get('doc_type', 'RUT')
        doc_number = data.get('doc_number', '')
        payer_email = data.get('payer_email')
        pedido_numero = data.get('pedido_numero')
        pedido_input = data.get('pedido_data')

        if not token or not payment_method_id:
            return Response(
                {"success": False, "message": "Faltan datos de la tarjeta (token o método de pago)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Extraer IP y User Agent del comprador
        ip_cliente = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip() or request.META.get('REMOTE_ADDR')
        user_agent = request.META.get('HTTP_USER_AGENT', '')

        # 1. Obtener o crear el pedido
        pedido = None
        if pedido_numero:
            pedido = Pedido.objects.filter(numero=pedido_numero).first()

        if not pedido and pedido_input:
            serializer = PedidoSerializer(data=pedido_input, context={'request': request})
            serializer.is_valid(raise_exception=True)
            pedido = serializer.save()

        if not pedido:
            return Response(
                {"success": False, "message": "No se pudo identificar el pedido para el cobro."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        is_test_card = data.get('is_test_card', False)
        card_last_digits = data.get('card_last_digits', '')

        # 2. Procesar el cobro directo en Mercado Pago API
        resultado = procesar_pago_directo_mercadopago(
            pedido=pedido,
            token=token,
            payment_method_id=payment_method_id,
            installments=installments,
            issuer_id=issuer_id,
            payer_email=payer_email,
            doc_type=doc_type,
            doc_number=doc_number,
            is_test_card=is_test_card,
            card_last_digits=card_last_digits,
            ip_cliente=ip_cliente,
            user_agent=user_agent,
        )

        resultado['pedido'] = PedidoSerializer(pedido).data
        http_code = status.HTTP_200_OK if resultado.get('success') else status.HTTP_400_BAD_REQUEST
        return Response(resultado, status=http_code)

    except Exception as exc:
        logger.error(f"Error en procesar_pago_tarjeta: {exc}")
        return Response(
            {"success": False, "message": f"Error al procesar el pago: {str(exc)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


class PanelEstadisticasView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from datetime import timedelta
        from .models import ItemPedido

        user = request.user
        if getattr(user, 'rol', '') != 'admin' and not user.is_staff and not user.is_superuser:
            return Response(
                {"detail": "No tienes permisos de administrador para consultar estadísticas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        periodo = request.query_params.get('periodo', 'todo')
        ahora = timezone.now()
        qs = Pedido.objects.all()

        if periodo == '7d':
            qs = qs.filter(creado_en__gte=ahora - timedelta(days=7))
        elif periodo == '30d':
            qs = qs.filter(creado_en__gte=ahora - timedelta(days=30))
        elif periodo == 'este_mes':
            qs = qs.filter(creado_en__year=ahora.year, creado_en__month=ahora.month)
        elif periodo == 'este_ano':
            qs = qs.filter(creado_en__year=ahora.year)

        # 1. KPIs Globales
        pedidos_pagados = qs.filter(estado__in=['pagado', 'en_proceso', 'enviado', 'entregado'])

        total_ventas_bruto = pedidos_pagados.aggregate(total=models.Sum('total'))['total'] or 0
        total_ventas_neto = pedidos_pagados.aggregate(neto=models.Sum('monto_neto'))['neto']
        if total_ventas_neto is None:
            total_comision = pedidos_pagados.aggregate(fee=models.Sum('comision_mp'))['fee'] or 0
            total_ventas_neto = max(0, total_ventas_bruto - float(total_comision))
        else:
            total_ventas_neto = float(total_ventas_neto)

        total_comision_mp = float(pedidos_pagados.aggregate(fee=models.Sum('comision_mp'))['fee'] or 0)
        total_pedidos_pagados_count = pedidos_pagados.count()
        total_pedidos_total_count = qs.count()

        ticket_promedio = round(total_ventas_bruto / total_pedidos_pagados_count) if total_pedidos_pagados_count > 0 else 0
        tasa_conversion = round((total_pedidos_pagados_count / total_pedidos_total_count) * 100, 1) if total_pedidos_total_count > 0 else 0

        conteo_estados = {
            'pagado': qs.filter(estado='pagado').count(),
            'en_proceso': qs.filter(estado='en_proceso').count(),
            'enviado': qs.filter(estado='enviado').count(),
            'entregado': qs.filter(estado='entregado').count(),
            'pendiente': qs.filter(estado='pendiente').count(),
            'cancelado': qs.filter(estado='cancelado').count(),
        }

        # 2. Top Productos Más Vendidos (Ordenado por Monto Total Recaudado y Unidades)
        items_pagados = ItemPedido.objects.filter(pedido__in=pedidos_pagados)
        top_prods_raw = (
            items_pagados.values('nombre', 'tipo', 'imagen')
            .annotate(
                unidades_vendidas=models.Sum('cantidad'),
                ingresos_totales=models.Sum(models.F('precio') * models.F('cantidad')),
            )
            .order_by('-ingresos_totales', '-unidades_vendidas')[:20]
        )

        top_productos = []
        for p in top_prods_raw:
            top_productos.append({
                'nombre': p['nombre'],
                'tipo': p['tipo'],
                'imagen': p['imagen'],
                'unidades_vendidas': p['unidades_vendidas'] or 0,
                'ingresos_totales': p['ingresos_totales'] or 0,
            })

        # 3. Ventas por Línea / Tipo de Producto
        lineas_raw = (
            items_pagados.values('linea')
            .annotate(
                unidades=models.Sum('cantidad'),
                ingresos=models.Sum(models.F('precio') * models.F('cantidad')),
            )
            .order_by('-ingresos')
        )
        ventas_por_linea = [
            {
                'linea': l['linea'] or 'Sin Línea / Personalizado',
                'unidades': l['unidades'] or 0,
                'ingresos': l['ingresos'] or 0,
            }
            for l in lineas_raw
        ]

        # 4. Distribución por Medio de Pago
        metodos_raw = (
            pedidos_pagados.values('metodo_pago', 'payment_method_id')
            .annotate(
                conteo=models.Count('id'),
                total=models.Sum('total'),
            )
            .order_by('-total')
        )
        medios_pago = [
            {
                'metodo': (m['payment_method_id'] or m['metodo_pago'] or 'mercadopago').upper(),
                'conteo': m['conteo'],
                'total': m['total'] or 0,
            }
            for m in metodos_raw
        ]

        # 5. Geografía / Regiones con Más Ventas
        regiones_raw = (
            pedidos_pagados.values('region')
            .annotate(
                conteo=models.Count('id'),
                total=models.Sum('total'),
            )
            .order_by('-total')[:6]
        )
        ventas_por_region = [
            {
                'region': r['region'] or 'Región Metropolitana',
                'conteo': r['conteo'],
                'total': r['total'] or 0,
            }
            for r in regiones_raw
        ]

        # 6. Listado de Pedidos del Período para la Tabla de Gestión (Todos los Estados)
        ultimas_transacciones = []
        for ped in qs.order_by('-creado_en')[:200]:
            ultimas_transacciones.append({
                'numero': ped.numero,
                'nombre': ped.nombre,
                'email': ped.email,
                'telefono': ped.telefono,
                'direccion': ped.direccion,
                'comuna': ped.comuna,
                'ciudad': ped.ciudad,
                'region': ped.region,
                'total': ped.total,
                'monto_neto': float(ped.monto_neto) if ped.monto_neto is not None else (ped.total - float(ped.comision_mp or 0)),
                'comision_mp': float(ped.comision_mp) if ped.comision_mp else 0,
                'estado': ped.estado,
                'metodo_pago': ped.metodo_pago,
                'payment_method_id': ped.payment_method_id,
                'card_last_four': ped.card_last_four,
                'card_first_six': ped.card_first_six,
                'authorization_code': ped.authorization_code,
                'ip_cliente': ped.ip_cliente,
                'historial_estados': ped.historial_estados or [],
                'pagado_en': ped.pagado_en,
                'creado_en': ped.creado_en,
            })

        return Response({
            'periodo': periodo,
            'kpis': {
                'total_ventas_bruto': total_ventas_bruto,
                'total_ventas_neto': total_ventas_neto,
                'total_comision_mp': total_comision_mp,
                'total_pedidos_pagados': total_pedidos_pagados_count,
                'total_pedidos_generados': total_pedidos_total_count,
                'ticket_promedio': ticket_promedio,
                'tasa_conversion': tasa_conversion,
                'conteo_estados': conteo_estados,
            },
            'top_productos': top_productos,
            'ventas_por_linea': ventas_por_linea,
            'medios_pago': medios_pago,
            'ventas_por_region': ventas_por_region,
            'ultimas_transacciones': ultimas_transacciones,
        }, status=status.HTTP_200_OK)


class ExportarVentasExcelView(generics.GenericAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        import io
        from datetime import timedelta
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        user = request.user
        if getattr(user, 'rol', '') != 'admin' and not user.is_staff and not user.is_superuser:
            return Response(
                {"detail": "Solo administradores pueden exportar el informe de ventas."},
                status=status.HTTP_403_FORBIDDEN
            )

        periodo = request.query_params.get('periodo', 'todo')
        ahora = timezone.now()
        qs = Pedido.objects.all()

        periodo_labels = {
            'todo': 'Histórico Completo (Todo el tiempo)',
            '7d': 'Últimos 7 días',
            '30d': 'Últimos 30 días',
            'este_mes': f'Mes Actual ({ahora.strftime("%B %Y")})',
            'este_ano': f'Año Actual ({ahora.year})',
        }

        if periodo == '7d':
            qs = qs.filter(creado_en__gte=ahora - timedelta(days=7))
        elif periodo == '30d':
            qs = qs.filter(creado_en__gte=ahora - timedelta(days=30))
        elif periodo == 'este_mes':
            qs = qs.filter(creado_en__year=ahora.year, creado_en__month=ahora.month)
        elif periodo == 'este_ano':
            qs = qs.filter(creado_en__year=ahora.year)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Paleta corporativa
        DARK_FILL = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
        ALT_ROW_FILL = PatternFill(start_color="F4F4F6", end_color="F4F4F6", fill_type="solid")
        WHITE_TEXT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        GOLD_TEXT_BOLD = Font(name="Arial", size=10.5, bold=True, color="9B7A3E")
        TITLE_FONT = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        REGULAR_FONT = Font(name="Arial", size=9.5, color="27272A")
        REGULAR_BOLD = Font(name="Arial", size=9.5, bold=True, color="27272A")

        THIN_BORDER = Border(
            left=Side(style='thin', color='E4E4E7'),
            right=Side(style='thin', color='E4E4E7'),
            top=Side(style='thin', color='E4E4E7'),
            bottom=Side(style='thin', color='E4E4E7'),
        )

        def style_header_row(ws, row_idx, max_col):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = DARK_FILL
                cell.font = WHITE_TEXT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER

        def auto_fit_columns(ws):
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # -------------------------------------------------------------
        # HOJA 1: RESUMEN EJECUTIVO & KPIS
        # -------------------------------------------------------------
        ws_resumen = wb.create_sheet(title="Resumen Ejecutivo & KPIs")
        ws_resumen.views.sheetView[0].showGridLines = True

        ws_resumen.merge_cells("A1:F2")
        title_cell = ws_resumen["A1"]
        title_cell.value = "RC ESTAMPA — INFORME EJECUTIVO DE VENTAS Y AUDITORÍA"
        title_cell.fill = DARK_FILL
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_resumen.cell(row=3, column=1, value=f"Período: {periodo_labels.get(periodo, periodo)}").font = GOLD_TEXT_BOLD
        ws_resumen.cell(row=3, column=4, value=f"Fecha de Emisión: {ahora.strftime('%d/%m/%Y %H:%M')}").font = REGULAR_FONT

        # KPIs Calc
        pedidos_pagados = qs.filter(estado__in=['pagado', 'en_proceso', 'enviado', 'entregado'])
        total_ventas_bruto = pedidos_pagados.aggregate(total=models.Sum('total'))['total'] or 0
        total_ventas_neto = pedidos_pagados.aggregate(neto=models.Sum('monto_neto'))['neto']
        if total_ventas_neto is None:
            total_comision = pedidos_pagados.aggregate(fee=models.Sum('comision_mp'))['fee'] or 0
            total_ventas_neto = max(0, total_ventas_bruto - float(total_comision))
        else:
            total_ventas_neto = float(total_ventas_neto)

        total_comision_mp = float(pedidos_pagados.aggregate(fee=models.Sum('comision_mp'))['fee'] or 0)
        total_pedidos_pagados_count = pedidos_pagados.count()
        total_pedidos_total_count = qs.count()
        ticket_promedio = round(total_ventas_bruto / total_pedidos_pagados_count) if total_pedidos_pagados_count > 0 else 0
        tasa_conversion = round((total_pedidos_pagados_count / total_pedidos_total_count) * 100, 1) if total_pedidos_total_count > 0 else 0

        # KPI Table
        ws_resumen.cell(row=5, column=1, value="MÉTRICA / KPI PRINCIPAL").font = WHITE_TEXT
        ws_resumen.cell(row=5, column=2, value="VALOR CONSOLIDADO").font = WHITE_TEXT
        ws_resumen.cell(row=5, column=3, value="DESCRIPCIÓN").font = WHITE_TEXT
        style_header_row(ws_resumen, 5, 3)

        kpi_rows = [
            ("Ventas Totales Brutas", total_ventas_bruto, "$#,##0", "Total bruto facturado y aprobado"),
            ("Ingresos Netos Reales", total_ventas_neto, "$#,##0", "Ingreso líquido tras descontar comisiones de pasarela"),
            ("Comisiones Retenidas Pasarela", total_comision_mp, "$#,##0", "Comisiones retenidas por Mercado Pago"),
            ("Total Pedidos Pagados / Exitosos", total_pedidos_pagados_count, "0", "Órdenes concretadas y aprobadas"),
            ("Total Pedidos Generados (Embudo)", total_pedidos_total_count, "0", "Total general de pedidos iniciados"),
            ("Ticket Promedio de Venta", ticket_promedio, "$#,##0", "Promedio recaudado por pedido pagado"),
            ("Tasa de Conversión / Aprobación", f"{tasa_conversion}%", "@", "Porcentaje de pedidos pagados vs creados"),
        ]

        for idx, (kpi_name, kpi_val, num_fmt, desc) in enumerate(kpi_rows, start=6):
            c1 = ws_resumen.cell(row=idx, column=1, value=kpi_name)
            c2 = ws_resumen.cell(row=idx, column=2, value=kpi_val)
            c3 = ws_resumen.cell(row=idx, column=3, value=desc)
            if num_fmt != "@":
                c2.number_format = num_fmt
            c1.font = REGULAR_BOLD
            c2.font = GOLD_TEXT_BOLD
            c3.font = REGULAR_FONT
            c1.border = c2.border = c3.border = THIN_BORDER
            if idx % 2 == 1:
                c1.fill = c2.fill = c3.fill = ALT_ROW_FILL

        # Top Productos
        start_row_top = 15
        ws_resumen.cell(row=start_row_top, column=1, value="TOP PRODUCTOS MÁS VENDIDOS").font = GOLD_TEXT_BOLD
        start_row_top += 1
        ws_resumen.cell(row=start_row_top, column=1, value="Producto").font = WHITE_TEXT
        ws_resumen.cell(row=start_row_top, column=2, value="Tipo").font = WHITE_TEXT
        ws_resumen.cell(row=start_row_top, column=3, value="Unidades").font = WHITE_TEXT
        ws_resumen.cell(row=start_row_top, column=4, value="Recaudación Total").font = WHITE_TEXT
        style_header_row(ws_resumen, start_row_top, 4)

        items_pagados = ItemPedido.objects.filter(pedido__in=pedidos_pagados)
        top_prods_raw = (
            items_pagados.values('nombre', 'tipo')
            .annotate(
                unidades=models.Sum('cantidad'),
                ingresos=models.Sum(models.F('precio') * models.F('cantidad')),
            )
            .order_by('-unidades')[:15]
        )

        for p_idx, prod in enumerate(top_prods_raw, start=start_row_top + 1):
            p1 = ws_resumen.cell(row=p_idx, column=1, value=prod['nombre'])
            p2 = ws_resumen.cell(row=p_idx, column=2, value=(prod['tipo'] or 'Catálogo').capitalize())
            p3 = ws_resumen.cell(row=p_idx, column=3, value=prod['unidades'])
            p4 = ws_resumen.cell(row=p_idx, column=4, value=prod['ingresos'])
            p4.number_format = "$#,##0"
            p1.font = REGULAR_FONT
            p2.font = REGULAR_FONT
            p3.font = REGULAR_BOLD
            p4.font = GOLD_TEXT_BOLD
            p1.border = p2.border = p3.border = p4.border = THIN_BORDER

        auto_fit_columns(ws_resumen)

        # -------------------------------------------------------------
        # FUNCIÓN AUXILIAR PARA POBLAR HOJAS DE DETALLE DE PEDIDOS
        # -------------------------------------------------------------
        COLUMNAS_DETALLE = [
            ("N° Pedido", 16),
            ("Fecha Creación", 18),
            ("Fecha Pago", 18),
            ("Cliente", 22),
            ("Email", 26),
            ("Teléfono", 16),
            ("Dirección", 28),
            ("Comuna", 18),
            ("Ciudad", 16),
            ("Región", 22),
            ("Estado", 14),
            ("Total Bruto ($)", 16),
            ("Comisión Pasarela ($)", 22),
            ("Monto Neto ($)", 16),
            ("Método Pago", 16),
            ("Medio / Franquicia", 18),
            ("Últimos 4 Dígitos", 16),
            ("BIN (6 Dígitos)", 15),
            ("Cód. Autorización", 18),
            ("Cuotas", 10),
            ("IP Cliente", 18),
        ]

        def poblar_hoja_pedidos(ws, pedidos_subset, titulo_hoja):
            ws.views.sheetView[0].showGridLines = True

            ws.merge_cells("A1:U1")
            h_cell = ws["A1"]
            h_cell.value = f"RC ESTAMPA — {titulo_hoja.upper()} ({periodo_labels.get(periodo, periodo)})"
            h_cell.fill = DARK_FILL
            h_cell.font = TITLE_FONT
            h_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            for col_i, (col_name, _) in enumerate(COLUMNAS_DETALLE, start=1):
                c = ws.cell(row=3, column=col_i, value=col_name)
                c.fill = DARK_FILL
                c.font = WHITE_TEXT
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = THIN_BORDER
            ws.row_dimensions[3].height = 24

            for r_idx, ped in enumerate(pedidos_subset.order_by('-creado_en'), start=4):
                fecha_creacion = ped.creado_en.strftime('%d/%m/%Y %H:%M') if ped.creado_en else ''
                fecha_pago = ped.pagado_en.strftime('%d/%m/%Y %H:%M') if ped.pagado_en else ''
                monto_neto_val = float(ped.monto_neto) if ped.monto_neto is not None else (ped.total - float(ped.comision_mp or 0))

                valores = [
                    ped.numero,
                    fecha_creacion,
                    fecha_pago,
                    ped.nombre,
                    ped.email,
                    ped.telefono,
                    ped.direccion,
                    ped.comuna,
                    ped.ciudad,
                    ped.region,
                    ped.get_estado_display() if hasattr(ped, 'get_estado_display') else ped.estado,
                    ped.total,
                    float(ped.comision_mp or 0),
                    monto_neto_val,
                    ped.metodo_pago,
                    ped.payment_method_id or ped.metodo_pago,
                    f"•••• {ped.card_last_four}" if ped.card_last_four else "-",
                    ped.card_first_six or "-",
                    ped.authorization_code or "-",
                    ped.cuotas or 1,
                    ped.ip_cliente or "-",
                ]

                for col_i, val in enumerate(valores, start=1):
                    cell = ws.cell(row=r_idx, column=col_i, value=val)
                    cell.font = REGULAR_FONT
                    cell.border = THIN_BORDER

                    if col_i in (12, 13, 14):
                        cell.number_format = "$#,##0"
                        cell.alignment = Alignment(horizontal="right")
                    elif col_i in (1, 11):
                        cell.font = REGULAR_BOLD
                        cell.alignment = Alignment(horizontal="center")
                    elif col_i in (2, 3, 17, 18, 19, 20, 21):
                        cell.alignment = Alignment(horizontal="center")

                    if r_idx % 2 == 1:
                        cell.fill = ALT_ROW_FILL

            for col_i, (_, col_w) in enumerate(COLUMNAS_DETALLE, start=1):
                col_letter = get_column_letter(col_i)
                ws.column_dimensions[col_letter].width = col_w

        # -------------------------------------------------------------
        # HOJAS SEPARADAS POR ESTADO
        # -------------------------------------------------------------
        ws_todos = wb.create_sheet(title="Todos los Pedidos")
        poblar_hoja_pedidos(ws_todos, qs, "Listado Maestro de Todos los Pedidos")

        ws_pagados = wb.create_sheet(title="Pagados & En Proceso")
        poblar_hoja_pedidos(ws_pagados, qs.filter(estado__in=['pagado', 'en_proceso']), "Pedidos Pagados y en Confección")

        ws_enviados = wb.create_sheet(title="Enviados (En Tránsito)")
        poblar_hoja_pedidos(ws_enviados, qs.filter(estado='enviado'), "Pedidos Despachados con Courier")

        ws_entregados = wb.create_sheet(title="Entregados")
        poblar_hoja_pedidos(ws_entregados, qs.filter(estado='entregado'), "Pedidos Entregados a Conformidad")

        ws_pendientes = wb.create_sheet(title="Pendientes & Cancelados")
        poblar_hoja_pedidos(ws_pendientes, qs.filter(estado__in=['pendiente', 'cancelado']), "Pedidos Pendientes de Pago o Cancelados")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"RC_Estampa_Ventas_{periodo}_{ahora.strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


