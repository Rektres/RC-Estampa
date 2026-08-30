import uuid
from pathlib import Path

from django.core.files.storage import default_storage
from django.db.models.deletion import ProtectedError
from rest_framework import mixins, status, viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from cuentas.permissions import IsAdminRol

from .filters import ProductoFilter, ProductoVajillaFilter
from .models import (
    Categoria,
    ColorEditor,
    FotoCliente,
    Linea,
    PrecioEditor,
    Producto,
    ProductoVajilla,
    Region,
    TallaStandard,
)
from .serializers import (
    CategoriaSerializer,
    ColorEditorSerializer,
    FotoClienteSerializer,
    ProductoSerializer,
    ProductoVajillaSerializer,
    ProductoVajillaWriteSerializer,
    ProductoWriteSerializer,
)


class ProductoViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    serializer_class = ProductoSerializer
    filterset_class = ProductoFilter
    ordering_fields = ['precio', 'creado_en', 'destacado']
    ordering = ['-creado_en']
    lookup_field = 'slug'

    def get_queryset(self):
        return (
            Producto.objects.filter(activo=True)
            .exclude(linea='sin_categoria')
            .select_related('categoria')
            .prefetch_related('variantes', 'imagenes')
            .distinct()
        )


class ProductoVajillaViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    serializer_class = ProductoVajillaSerializer
    filterset_class = ProductoVajillaFilter
    ordering_fields = ['precio', 'creado_en', 'destacado']
    ordering = ['-creado_en']
    lookup_field = 'slug'

    def get_queryset(self):
        return (
            ProductoVajilla.objects.filter(activo=True)
            .exclude(linea='sin_categoria')
            .select_related('categoria')
            .prefetch_related('variantes', 'imagenes')
            .distinct()
        )


class CategoriaViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    pagination_class = None


class FotoClienteViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = FotoCliente.objects.all()
    serializer_class = FotoClienteSerializer
    pagination_class = None


# ---------------------------------------------------------------------------
# Panel de administración (CRUD completo, solo rol admin)
# ---------------------------------------------------------------------------

class PanelProductoViewSet(viewsets.ModelViewSet):
    """CRUD de productos de ropa. Incluye inactivos (a diferencia del público)."""

    serializer_class = ProductoWriteSerializer
    permission_classes = [IsAdminRol]
    filterset_class = ProductoFilter
    ordering_fields = ['precio', 'creado_en', 'destacado']
    ordering = ['-creado_en']

    def get_queryset(self):
        return (
            Producto.objects.all()
            .select_related('categoria')
            .prefetch_related('variantes', 'imagenes')
            .distinct()
        )


class PanelProductoVajillaViewSet(viewsets.ModelViewSet):
    serializer_class = ProductoVajillaWriteSerializer
    permission_classes = [IsAdminRol]
    filterset_class = ProductoVajillaFilter
    ordering_fields = ['precio', 'creado_en', 'destacado']
    ordering = ['-creado_en']

    def get_queryset(self):
        return (
            ProductoVajilla.objects.all()
            .select_related('categoria')
            .prefetch_related('variantes', 'imagenes')
            .distinct()
        )


class PanelCategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [IsAdminRol]
    pagination_class = None

    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except ProtectedError:
            return Response(
                {'detail': 'La categoría tiene productos asociados; reasígnalos o elimínalos primero.'},
                status=status.HTTP_400_BAD_REQUEST,
            )


class PanelUploadView(APIView):
    """Sube una imagen de producto y devuelve su URL en /media/."""

    permission_classes = [IsAdminRol]
    parser_classes = [MultiPartParser]

    MAX_SIZE = 5 * 1024 * 1024
    ALLOWED_EXT = {'.jpg', '.jpeg', '.png', '.webp', '.svg', '.gif'}

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Falta el archivo (campo "file").'}, status=400)
        if file.size > self.MAX_SIZE:
            return Response({'detail': 'La imagen supera el máximo de 5MB.'}, status=400)
        ext = Path(file.name).suffix.lower()
        if ext not in self.ALLOWED_EXT:
            return Response({'detail': f'Extensión no permitida: {ext}'}, status=400)
        path = default_storage.save(f'productos/{uuid.uuid4().hex}{ext}', file)
        return Response({'url': f'/media/{path}'}, status=201)


class PanelLineasView(APIView):
    """Gestión completa de Líneas / Colecciones de productos con persistencia en BD."""

    permission_classes = [IsAdminRol]

    def get(self, request):
        # 1. Inicializar líneas estándar si la tabla está vacía
        if Linea.objects.count() == 0:
            Linea.objects.create(nombre='Urbana', slug='urbana')
            Linea.objects.create(nombre='Formal', slug='formal')
            Linea.objects.create(nombre='Drinkware', slug='drinkware')

        # 2. Asegurar siempre la línea 'Ropa sin categoría' para prendas desasignadas
        Linea.objects.get_or_create(
            slug='sin_categoria',
            defaults={'nombre': 'Ropa sin categoría', 'es_sin_categoria': True}
        )

        # 3. Sincronizar cualquier línea existente en productos/categorías que no esté en la tabla
        lineas_usadas = set(
            list(Producto.objects.values_list('linea', flat=True).distinct()) +
            list(ProductoVajilla.objects.values_list('linea', flat=True).distinct()) +
            list(Categoria.objects.values_list('linea', flat=True).distinct())
        )
        for slug_usado in lineas_usadas:
            if slug_usado and not Linea.objects.filter(slug=slug_usado).exists():
                Linea.objects.create(
                    slug=slug_usado,
                    nombre=slug_usado.replace('_', ' ').replace('-', ' ').title(),
                    es_sin_categoria=(slug_usado == 'sin_categoria')
                )

        lineas = Linea.objects.all().order_by('es_sin_categoria', 'nombre')
        datos = []
        for l in lineas:
            count_ropa = Producto.objects.filter(linea=l.slug).count()
            count_vajilla = ProductoVajilla.objects.filter(linea=l.slug).count()
            count_cats = Categoria.objects.filter(linea=l.slug).count()
            datos.append({
                'id': l.id,
                'linea': l.slug,
                'nombre': l.nombre,
                'es_sin_categoria': l.es_sin_categoria,
                'total_productos': count_ropa + count_vajilla,
                'total_ropa': count_ropa,
                'total_drinkware': count_vajilla,
                'total_categorias': count_cats,
            })
        return Response(datos)

    def post(self, request):
        old_linea = request.data.get('old_linea', '').strip().lower()
        new_linea = request.data.get('new_linea', '').strip().lower()
        nombre = request.data.get('nombre', '').strip()

        if not new_linea:
            return Response({'detail': 'El identificador de la línea es obligatorio.'}, status=400)

        if not nombre:
            nombre = new_linea.replace('_', ' ').replace('-', ' ').title()

        if old_linea:
            linea_obj = Linea.objects.filter(slug=old_linea).first()
            if linea_obj:
                linea_obj.slug = new_linea
                linea_obj.nombre = nombre
                linea_obj.save()
            else:
                Linea.objects.get_or_create(slug=new_linea, defaults={'nombre': nombre})

            # Actualizar referencias en productos y categorías si cambió el slug
            if old_linea != new_linea:
                Producto.objects.filter(linea=old_linea).update(linea=new_linea)
                ProductoVajilla.objects.filter(linea=old_linea).update(linea=new_linea)
                Categoria.objects.filter(linea=old_linea).update(linea=new_linea)
        else:
            # Crear nueva línea en base de datos
            Linea.objects.get_or_create(slug=new_linea, defaults={'nombre': nombre})

        return Response({'success': True, 'linea': new_linea, 'nombre': nombre}, status=status.HTTP_200_OK)

    def delete(self, request):
        linea = request.query_params.get('linea', '').strip().lower()
        reassign_to = request.query_params.get('reassign_to', '').strip().lower() or 'sin_categoria'

        if not linea:
            return Response({'detail': 'Debes especificar la línea a eliminar.'}, status=400)

        # Regla del sistema: Siempre debe existir al menos 1 línea activa
        lineas_activas_restantes = Linea.objects.filter(es_sin_categoria=False).exclude(slug=linea).count()
        if lineas_activas_restantes < 1:
            return Response({
                'detail': 'No puedes eliminar la última línea activa. Como regla del sistema, siempre debe existir al menos 1 línea disponible.'
            }, status=400)

        # Asegurar que la línea 'sin_categoria' exista en la BD
        Linea.objects.get_or_create(
            slug='sin_categoria',
            defaults={'nombre': 'Ropa sin categoría', 'es_sin_categoria': True}
        )

        # Traspasar todos los productos y categorías asociados
        Producto.objects.filter(linea=linea).update(linea=reassign_to)
        ProductoVajilla.objects.filter(linea=linea).update(linea=reassign_to)
        Categoria.objects.filter(linea=linea).update(linea=reassign_to)

        # Eliminar el registro de la línea
        Linea.objects.filter(slug=linea).delete()

        return Response({'success': True, 'reassigned_to': reassign_to})


@api_view(['GET'])
@permission_classes([AllowAny])
def editor_config(request):
    """Colores, precios base y tallas para el editor de diseño."""
    colores = ColorEditorSerializer(ColorEditor.objects.all(), many=True).data
    precios = {p.producto_key: p.precio for p in PrecioEditor.objects.all()}
    tallas = list(TallaStandard.objects.values_list('nombre', flat=True))
    regiones = list(Region.objects.values_list('nombre', flat=True))
    return Response({
        'colores': colores,
        'precios': precios,
        'tallas': tallas,
        'regiones': regiones,
    })


class PanelExportarProductosExcelView(APIView):
    """Exporta catálogo completo de Ropa o Drinkware a Excel con colores corporativos y hojas por categoría."""

    permission_classes = [IsAdminRol]

    def get(self, request):
        import io
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from django.http import HttpResponse
        from django.utils import timezone

        tipo = request.query_params.get('tipo', 'ropa')
        ahora = timezone.now()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Quitar hoja default vacía

        header_fill = PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0'),
        )

        if tipo == 'ropa':
            headers = [
                'ID', 'Nombre del Producto', 'Línea', 'Categoría',
                'Precio Normal (CLP)', 'Precio Oferta (CLP)', 'Stock Total',
                'Variantes (Talla / Color / Stock / SKU)',
                'Estado', 'Destacado', 'Nuevo', 'Slug', 'Creado En', 'Foto Principal'
            ]
            productos = (
                Producto.objects.select_related('categoria')
                .prefetch_related('variantes', 'imagenes')
                .all()
                .order_by('categoria__nombre', 'nombre')
            )
            cats_dict = {}
            for p in productos:
                cat_name = p.categoria.nombre if p.categoria else 'Sin Categoría'
                cats_dict.setdefault(cat_name, []).append(p)

            self._write_sheet(wb, "Todo Textil", headers, productos, 'ropa', header_fill, header_font, regular_font, thin_border)
            for cat_name, prods in cats_dict.items():
                safe_title = cat_name[:30].replace('/', '-').replace('\\', '-')
                self._write_sheet(wb, safe_title, headers, prods, 'ropa', header_fill, header_font, regular_font, thin_border)

            filename = f"RC_Estampa_Catalogo_Ropa_{ahora.strftime('%Y%m%d_%H%M')}.xlsx"

        else:
            headers = [
                'ID', 'Nombre del Producto', 'Línea', 'Categoría',
                'Material', 'Capacidad (ml)',
                'Precio Normal (CLP)', 'Precio Oferta (CLP)', 'Stock Total',
                'Variantes (Color / Stock / SKU)',
                'Estado', 'Destacado', 'Nuevo', 'Slug', 'Creado En', 'Foto Principal'
            ]
            productos = (
                ProductoVajilla.objects.select_related('categoria')
                .prefetch_related('variantes', 'imagenes')
                .all()
                .order_by('categoria__nombre', 'nombre')
            )
            cats_dict = {}
            for p in productos:
                cat_name = p.categoria.nombre if p.categoria else 'Sin Categoría'
                cats_dict.setdefault(cat_name, []).append(p)

            self._write_sheet(wb, "Todo Drinkware", headers, productos, 'drinkware', header_fill, header_font, regular_font, thin_border)
            for cat_name, prods in cats_dict.items():
                safe_title = cat_name[:30].replace('/', '-').replace('\\', '-')
                self._write_sheet(wb, safe_title, headers, prods, 'drinkware', header_fill, header_font, regular_font, thin_border)

            filename = f"RC_Estampa_Catalogo_Drinkware_{ahora.strftime('%Y%m%d_%H%M')}.xlsx"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response

    def _write_sheet(self, wb, title, headers, productos, tipo, header_fill, header_font, regular_font, thin_border):
        import openpyxl
        from openpyxl.styles import Alignment

        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for p in productos:
            variantes = list(p.variantes.all())
            stock_total = sum(v.stock for v in variantes)
            if tipo == 'ropa':
                var_str = "; ".join(f"Talla: {v.talla}, Color: {v.color}, Stock: {v.stock}, SKU: {v.sku}" for v in variantes)
            else:
                var_str = "; ".join(f"Color: {v.color}, Stock: {v.stock}, SKU: {v.sku}" for v in variantes)

            foto = p.imagenes.first().imagen if p.imagenes.exists() else ''

            if tipo == 'ropa':
                row = [
                    p.id,
                    p.nombre,
                    p.linea.upper(),
                    p.categoria.nombre if p.categoria else 'Sin Categoría',
                    p.precio,
                    p.precio_oferta if p.precio_oferta else '',
                    stock_total,
                    var_str,
                    'ACTIVO' if p.activo else 'DESHABILITADO',
                    'SÍ' if p.destacado else 'NO',
                    'SÍ' if p.nuevo else 'NO',
                    p.slug,
                    p.creado_en.strftime('%Y-%m-%d %H:%M') if p.creado_en else '',
                    foto,
                ]
            else:
                row = [
                    p.id,
                    p.nombre,
                    p.linea.upper(),
                    p.categoria.nombre if p.categoria else 'Sin Categoría',
                    p.material or '',
                    p.capacidad_ml if p.capacidad_ml else '',
                    p.precio,
                    p.precio_oferta if p.precio_oferta else '',
                    stock_total,
                    var_str,
                    'ACTIVO' if p.activo else 'DESHABILITADO',
                    'SÍ' if p.destacado else 'NO',
                    'SÍ' if p.nuevo else 'NO',
                    p.slug,
                    p.creado_en.strftime('%Y-%m-%d %H:%M') if p.creado_en else '',
                    foto,
                ]
            ws.append(row)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.font = regular_font
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)
